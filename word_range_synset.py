from all_synsets_get import All_Synstes_Get
# Wordnetをインポート
from nltk.corpus import wordnet as wn
from neo4j import GraphDatabase
import neo4j_write_synset as write_neo4j
import matplotlib.pyplot as plt
import networkx as nx
import openpyxl

driver = GraphDatabase.driver('bolt://localhost:7687', auth=('neo4j', 'ikkyu2419'))

class Word_Range_Synset(All_Synstes_Get):
    def __init__(self):
        super().__init__()
        self.jpn_word_range_dict = {}
        self.eng_word_range_dict = {}
        self.compare_list = []
        self.parents_different_list = []
        self.brother_different_list = []
        self.eng_target = []
        self.word_graph = nx.Graph()

    # 隣接ノードを取得する
    def neighbors_get(self, start):
        print(self.G.neighbors(start))
        return next(self.G.neighbors(start))

    # 単語の頻出範囲を調べる
    def range_get(self, target, lang):
        jpn_dict = {}
        eng_dict = {}
        # 概念に紐づいた単語群を取得
        lemmas = self.word_get(target, lang)
        # 単語を一つずつ見て、上位概念と下位概念に存在するかを確認
        for lemma in lemmas:
            range_list = [target]
            # 兄弟を探索
            range_list = self.search_brother_word(lemma, target, range_list, lang)

            # 上位概念を探索
            range_list = self.search_hypernym_word(lemma, target, range_list, lang)

            # 下位概念を探索
            range_list = self.search_hyponym_word(lemma, target, range_list, lang)

            if lang == 'jpn':
                self.jpn_word_range_dict[lemma] = len(range_list)
                jpn_dict[lemma] = range_list
            elif lang == 'eng':
                self.eng_word_range_dict[lemma] = len(range_list)

                eng_dict[lemma] = range_list
            print('単語"{}"の範囲は{}'.format(lemma, len(range_list)))
        if lang=='jpn':
            return jpn_dict
        elif lang == 'eng':
            return eng_dict

    # 単語の頻出範囲を調べる
    def parents_range_get(self, target, lang):
        jpn_dict = {}
        eng_dict = {}
        # 概念に紐づいた単語群を取得
        lemmas = self.word_get(target, lang)
        # 単語を一つずつ見て、上位概念と下位概念に存在するかを確認
        for lemma in lemmas:
            range_list = [target]

            # 上位概念を探索
            range_list = self.search_hypernym_word(lemma, target, range_list, lang)

            # 下位概念を探索
            range_list = self.search_hyponym_word(lemma, target, range_list, lang)

            if lang == 'jpn':
                self.jpn_word_range_dict[lemma] = len(range_list)
                jpn_dict[lemma] = range_list
            elif lang == 'eng':
                self.eng_word_range_dict[lemma] = len(range_list)

                eng_dict[lemma] = range_list
            print('単語"{}"の範囲は{}'.format(lemma, len(range_list)))
        if lang=='jpn':
            return jpn_dict
        elif lang == 'eng':
            return eng_dict

    # 単語の頻出範囲を調べる
    def brother_range_get(self, target, lang):
        jpn_dict = {}
        eng_dict = {}
        # 概念に紐づいた単語群を取得
        lemmas = self.word_get(target, lang)
        # 単語を一つずつ見て、上位概念と下位概念に存在するかを確認
        for lemma in lemmas:
            range_list = [target]
            # 兄弟を探索
            range_list = self.search_brother_word(lemma, target, range_list, lang)

            if lang == 'jpn':
                self.jpn_word_range_dict[lemma] = len(range_list)
                jpn_dict[lemma] = range_list
            elif lang == 'eng':
                self.eng_word_range_dict[lemma] = len(range_list)

                eng_dict[lemma] = range_list
            print('単語"{}"の範囲は{}'.format(lemma, len(range_list)))
        if lang=='jpn':
            return jpn_dict
        elif lang == 'eng':
            return eng_dict

    # 単語同士で範囲に違いがある単語の組み合わせをリストに追加
    def word_compare(self, synset):
        jpn_dict = self.range_get(synset, 'jpn')
        print(jpn_dict)
        eng_dict = self.range_get(synset, 'eng')

        for item_jpn in jpn_dict.items():
            for item_eng in eng_dict.items():
                if set(item_jpn[1]) != set(item_eng[1]):
                    if [item_jpn[0], item_eng[0], item_jpn[1], item_eng[1]] not in self.compare_list:
                        #self.word_graph.add_edge(item_jpn[0], item_eng[0])
                        self.compare_list.append([item_jpn[0], item_eng[0], item_jpn[1], item_eng[1]])
                        if len(item_jpn[1]) < len(item_eng[1]):
                            self.eng_target.append([item_jpn[0], item_eng[0], item_jpn[1], item_eng[1]])

    # 単語同士で範囲に違いがある単語の組み合わせをリストに追加
    def parents_brother_word_compare(self, synset):

        # 日本語の単語とその範囲を取り出す
        for item_jpn in self.parents_range_get(synset, 'jpn').items():
            # 英語の単語とその範囲を取り出す
            for item_eng in self.parents_range_get(synset, 'eng').items():
                # 概念の範囲が異なるかを確認
                if set(item_jpn[1]) != set(item_eng[1]):
                    # 異なった単語の組み合わせと同一の単語の組み合わせが入っていないかを確認する
                    for parents_records in self.parents_different_list:
                        if item_jpn[0] == parents_records[0] and item_eng[0] == parents_records[1]:
                            break
                    else:
                        self.parents_different_list.append([item_jpn[0], item_eng[0], item_jpn[1], item_eng[1]])


        for item_jpn in self.brother_range_get(synset, 'jpn').items():
            for item_eng in self.brother_range_get(synset, 'eng').items():
                if set(item_jpn[1]) != set(item_eng[1]):
                    for brother_records in self.brother_different_list:
                        if item_jpn[0] == brother_records[0] and item_eng[0] == brother_records[1]:
                            break
                    else:
                        self.brother_different_list.append([item_jpn[0], item_eng[0], item_jpn[1], item_eng[1]])

    def print_compare_result(self, target_word = None):
        lists = [self.parents_different_list, self.brother_different_list]
        for list in lists:
            for data in list:
                if target_word == None:
                    print(data)
                else:
                    if target_word in data:
                        print(data)
        print('親子は{}'.format(len(lists[0])))
        print('兄弟は{}'.format(len(lists[1])))

    def write_excel(self):
        for list in [self.parents_different_list, self.brother_different_list]:
            book = openpyxl.Workbook()
            sheet = book.active
            row = 1
            for data in list:
                max_row = sheet.max_row
                sheet.cell(row = max_row+1, column = 1).value = data[0]
                sheet.cell(row = max_row+1, column = 2).value = data[1]
                sheet.cell(row = max_row+1, column = 3).value = str(data[2])
                sheet.cell(row = max_row+1, column = 4).value = str(data[3])
                sheet.cell(row = max_row+1, column = 5).value = len(data[2])
                sheet.cell(row = max_row+1, column = 6).value = len(data[3])
            if list == self.parents_different_list:
                book.save('../../研究/WordNet比較データ/親子_所属単語比較_1215.xlsx')
            elif list == self.brother_different_list:
                book.save('../../研究/WordNet比較データ/兄弟_所属単語比較_1215.xlsx')

    def search_brother_word(self, lemma, target, range_list, lang):
        brothers = self.get_brothers(target)
        if brothers != []:
            for brother in brothers:
                words = self.word_get(brother, lang)
                if lemma in words:
                    range_list.append(brother)

        return range_list

    def search_hypernym_word(self, lemma, target, range_list, lang):
        hypernyms = self.get_hypernyms(target)
        for hypernym in hypernyms:
            words = self.word_get(hypernym, lang)
            if lemma in words:
                range_list.append(hypernym)
                range_list = self.search_hypernym_word(lemma, hypernym, range_list, lang)

        return range_list

    # 下位概念に単語が含まれていないか確認する
    def search_hyponym_word(self, lemma, target, range_list, lang):
        hyponyms = self.get_hyponyms(target)
        if hyponyms != []:
            for hyponym in hyponyms:
                words = self.word_get(hyponym, lang)
                if lemma in words:
                    range_list.append(hyponym)
                    range_list = self.search_hyponym_word(lemma, hyponym, range_list, lang)

        return range_list

    # 概念から単語を取得する
    def word_get(self, target, lang):
        return target.lemma_names(lang)

    def link_draw(self):
        nx.draw(self.word_graph, with_labels=True)
        plt.show()

if __name__ == '__main__':
    word_range = Word_Range_Synset()
    word_range.search_DFS(wn.synsets('entity')[0])
    bfs_list = [wn.synsets('entity')[0]]
    exist_list = []
    
    # 幅優先探索で一つずつノードを取り出す
    for bfs_edge_link_pare in list(nx.bfs_edges(word_range.G, source=bfs_list[0])):
        word_range.parents_brother_word_compare(bfs_edge_link_pare[1])

    word_range.print_compare_result()
    word_range.write_excel()

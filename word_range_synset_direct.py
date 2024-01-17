from all_synsets_get_direct import All_Synsets_Get_Direct
from nltk.corpus import wordnet as wn
import networkx as nx
import openpyxl

class Word_Range_Synset(All_Synsets_Get_Direct):
    def __init__(self):
        super().__init__()
        self.first_lang = 'cmn'
        self.second_lang = 'ind'
        self.eng_word_range_dict = {}
        self.jpn_word_range_dict = {}
        self.cmn_word_range_dict = {}
        self.ind_word_range_dict = {}
        self.compare_list = []
        self.different_list = []
        self.parents_different_list = []
        self.brother_different_list = []
        self.eng_target = []
        self.common_synset_per_list = []

    def search_brother_word(self, lemma, synset, range_list, lang):
        brothers = self.get_brothers(synset)
        if brothers != []:
            for brother in brothers:
                words = self.word_get(brother, lang)
                if lemma in words:
                    range_list.append(brother)

        return range_list

    def search_hypernym_word(self, lemma, synset, range_list, lang):
        hypernyms = self.get_parent(synset)
        for hypernym in hypernyms:
            words = self.word_get(hypernym, lang)
            # 単語が空だった場合、上位の概念から単語群を取得
            if words == []:
                for more_hypernym in self.get_parent(hypernym):
                    words.extend(self.word_get(more_hypernym, lang))
            if lemma in words:
                range_list.append(hypernym)
                range_list = self.search_hypernym_word(lemma, hypernym, range_list, lang)
        
        #print('hypernyms:{}, target:{}, lemma:{}, lang:{}'.format(hypernyms,target, lemma ,lang))
        return range_list

    # 下位概念に単語が含まれていないか確認する
    def search_hyponym_word(self, lemma, synset, range_list, lang):
        hyponyms = self.get_children(synset)
        if hyponyms != []:
            for hyponym in hyponyms:
                words = self.word_get(hyponym, lang)
                # 単語が対応づいていない時、さらに下の概念から単語を借りてくる
                if words == []:
                    words.extend(self.word_get(synset, lang))
                if lemma in words:
                    range_list.append(hyponym)
                    range_list = self.search_hyponym_word(lemma, hyponym, range_list, lang)

        return range_list
    
    def search_synset_word(self, lemma, synset, range_list, lang, visited_list):
        #print(synset)
        # 探索するsynsetのリスト
        search_list = self.get_children(synset) + self.get_parent(synset) + self.get_brothers(synset)
        # 探索済みのsynsetを省く
        search_list = [search for search in list(set(search_list)) if search not in list(set(range_list + visited_list))]
        

        if search_list != []:
            visited_list += search_list
            for search_synset in search_list:
                words = self.word_get(search_synset, lang)
                # 単語がない場合, 上位から取得する
                """if words == []:
                    for hypernym in self.get_parent(search_synset):
                        words.extend(self.word_get(hypernym, lang))"""
                if lemma in words:
                    range_list.append(search_synset)
                    range_list, visited_list = self.search_synset_word(lemma, search_synset, range_list, lang, visited_list)
        
        return range_list, visited_list


    # synsetの単語群を返すメソッド
    def word_get(self, synset, lang):
        return synset.lemma_names(lang)
    
    def word_range_get(self, synset, language):
        # 単語の範囲を入れる辞書
        word_dict = {}
        # 単語群を取得
        lemmas = self.word_get(synset, language)

        # 単語ごとに概念範囲を取得していく
        for lemma in lemmas:
            range_list = [synset]
            visited_list = [synset]
            range_list, visited_list = self.search_synset_word(lemma, synset, range_list, language, visited_list)

            if language == 'jpn':
                self.jpn_word_range_dict[lemma] = len(range_list)
            elif language == 'cmn':
                self.cmn_word_range_dict[lemma] = len(range_list)
            elif language == 'eng':
                self.eng_word_range_dict[lemma] = len(range_list)
            elif language == 'ind':
                self.ind_word_range_dict[lemma] = len(range_list)
            word_dict[lemma] = range_list
        return word_dict

    def parent_range_get(self, synset, language):
        # 単語の範囲を入れる辞書
        word_dict = {}
        # 単語群を取得
        lemmas = self.word_get(synset, language)

        for lemma in lemmas:
            range_list = [synset]
            range_list = self.search_hypernym_word(lemma, synset, range_list, language)
            range_list = self.search_hyponym_word(lemma, synset, range_list, language)

            if language == 'jpn':
                self.jpn_word_range_dict[lemma] = len(range_list)
            elif language == 'cmn':
                self.cmn_word_range_dict[lemma] = len(range_list)
            elif language == 'eng':
                self.eng_word_range_dict[lemma] = len(range_list)
            elif language == 'ind':
                self.ind_word_range_dict[lemma] = len(range_list)
            word_dict[lemma] = range_list
        return word_dict

        # 単語の頻出範囲を調べる
    def brother_range_get(self, synset, language):
        word_dict = {}
        # 概念に紐づいた単語群を取得
        lemmas = self.word_get(synset, language)
        # 単語を一つずつ見て、上位概念と下位概念に存在するかを確認
        for lemma in lemmas:
            range_list = [synset]
            # 兄弟を探索
            range_list = self.search_brother_word(lemma, synset, range_list, language)

            if language == 'jpn':
                self.jpn_word_range_dict[lemma] = len(range_list)
            elif language == 'eng':
                self.eng_word_range_dict[lemma] = len(range_list)
            elif language == 'cmn':
                self.cmn_word_range_dict[lemma] = len(range_list)
            elif language == 'ind':
                self.ind_word_range_dict[lemma] = len(range_list)
            word_dict[lemma] = range_list

        return word_dict

    # 単語の概念範囲の比較
    def word_compare_split_par_bro(self, synset):
        language_list = [self.first_lang, self.second_lang]
        for language in language_list:
            # 日本語の単語とその範囲を取り出す
            for item in self.parent_range_get(synset, language).items():
                lang = [language]
                # 単語が所属している概念を一つずつ取り出す
                for including_synset in item[1]:
                    jpn_and_cmn_range = 0
                    # 英語の単語とその範囲を取り出す
                    for compare_item in self.parent_range_get(including_synset, list(set(language_list)^set(lang))[0]).items():
                        #print('item:{}, compare_item:{}, type:{}'.format(item[1], compare_item[1], type(compare_item[1][0])))
                        commmon_synset_num = len(set(item[1])&set(compare_item[1]))
                        #print((commmon_synset_num/(len(item[1])+len(compare_item[1])-commmon_synset_num))*100)
                        self.common_synset_per_list.append((commmon_synset_num/(len(item[1])+len(compare_item[1])-commmon_synset_num))*100)
                        # 概念の範囲が異なるかを確認
                        if set(item[1]) != set(compare_item[1]): # 単語が付随しているsynsetが異なる時
                            if language == self.first_lang:
                                # 異なった単語の組み合わせと同一の単語の組み合わせが入っていないかを確認する
                                for records in self.parents_different_list:
                                    if item[0] == records[0] and compare_item[0] == records[1]:
                                        break
                                else:
                                    self.parents_different_list.append([item[0], compare_item[0], item[1], compare_item[1]])
                            elif language == self.second_lang:
                                # 異なった単語の組み合わせと同一の単語の組み合わせが入っていないかを確認する
                                for records in self.parents_different_list:
                                    if compare_item[0] == records[0] and item[0] == records[1]:
                                        break
                                else:
                                    self.parents_different_list.append([compare_item[0], item[0], compare_item[1], item[1]])

            #兄弟の単語で比較
            for item_jpn in self.brother_range_get(synset, self.first_lang).items():
                for item_eng in self.brother_range_get(synset, self.second_lang).items():
                    if set(item_jpn[1]) != set(item_eng[1]):
                        for brother_records in self.brother_different_list:
                            if item_jpn[0] == brother_records[0] and item_eng[0] == brother_records[1]:
                                break
                        else:
                            self.brother_different_list.append([item_jpn[0], item_eng[0], item_jpn[1], item_eng[1]])


    # 単語の概念範囲の比較
    def word_compare(self, synset):
        language_list = [self.first_lang, self.second_lang]
        for language in language_list:
            #単語で比較
            for item in self.word_range_get(synset, language).items():
                lang = [language]
                for including_synset in item[1]:
                    for compare_item in self.word_range_get(including_synset, list(set(language_list)^set(lang))[0]).items():
                        if set(item[1]) != set(compare_item[1]): # 単語が付随しているsynsetが異なる時
                            if language == self.first_lang:
                                # 異なった単語の組み合わせと同一の単語の組み合わせが入っていないかを確認する
                                for records in self.different_list:
                                    if item[0] == records[0] and compare_item[0] == records[1]:
                                        break
                                else:
                                    self.different_list.append([item[0], compare_item[0], item[1], compare_item[1]])
                            elif language == self.second_lang:
                                # 異なった単語の組み合わせと同一の単語の組み合わせが入っていないかを確認する
                                for records in self.different_list:
                                    if compare_item[0] == records[0] and item[0] == records[1]:
                                        break
                                else:
                                    self.different_list.append([compare_item[0], item[0], compare_item[1], item[1]])

    def word_confirm(self, item_synset, compare_synset):
        for item in item_synset:
            if item.lemma_names(self.first_lang) == [] or item.lemma_names(self.second_lang) == []:
                return False
        for item2 in compare_synset:
            if item2.lemma_names(self.first_lang) == [] or item2.lemma_names(self.second_lang) == []:
                return False
        return True

    def print_compare_result(self, target_word = None):
        #lists = [self.parents_different_list, self.brother_different_list] #
        lists = [self.different_list] #


        """for list in lists:
            for data in list:
                if target_word == None:
                    print(data)
                else:
                    if target_word in data:
                        print(data)"""
        print('親子は{}'.format(len(lists[0])))
        #print('兄弟は{}'.format(len(lists[1])))
        print(target_word)

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 1
        for data in self.different_list:
            #print(data[0])
            max_row = sheet.max_row
            sheet.cell(row = max_row+1, column = 1).value = data[0]
            sheet.cell(row = max_row+1, column = 2).value = data[1]
            sheet.cell(row = max_row+1, column = 3).value = str(data[2])
            sheet.cell(row = max_row+1, column = 4).value = str(data[3])
            sheet.cell(row = max_row+1, column = 5).value = len(data[2])
            sheet.cell(row = max_row+1, column = 6).value = len(data[3])
        book.save('../../../../ikkyu/Documents/研究/WordNet比較データ/単語ペア取得_1211_中国尼.xlsx')

        """for list in [self.parents_different_list, self.brother_different_list]:
            book = openpyxl.Workbook()
            sheet = book.active
            row = 1
            for data in list:
                print(data[0])
                max_row = sheet.max_row
                sheet.cell(row = max_row+1, column = 1).value = data[0]
                sheet.cell(row = max_row+1, column = 2).value = data[1]
                sheet.cell(row = max_row+1, column = 3).value = str(data[2])
                sheet.cell(row = max_row+1, column = 4).value = str(data[3])
                sheet.cell(row = max_row+1, column = 5).value = len(data[2])
                sheet.cell(row = max_row+1, column = 6).value = len(data[3])
            if list == self.parents_different_list:
                book.save('../../../../ikkyu/Documents/研究/WordNet比較データ/親子_所属単語比較_1018_日本インドネシア.xlsx')
            elif list == self.brother_different_list:
                book.save('../../../../ikkyu/Documents/研究/WordNet比較データ/兄弟_所属単語比較_1018_日本インドネシア.xlsx')"""

if __name__ == '__main__':
    word_range = Word_Range_Synset()
    word_range.direct_graph_make_DFS(wn.synsets('entity')[0])
    bfs_list = [wn.synsets('entity')[0]]

    for bfs_edge_link_pare in list(nx.bfs_edges(word_range.G, source=bfs_list[0])):
        print(bfs_edge_link_pare[1])
        word_range.word_compare(bfs_edge_link_pare[1])

    word_range.print_compare_result()
    word_range.write_excel()
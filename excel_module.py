"""
2023/6/22, エクセルファイルを読み込んでlistに入れる関数
"""

import openpyxl
import os

def excel_read_getter(filepath, sheetname='Sheet1', min_row=0, min_col=0):
    book = openpyxl.load_workbook(filepath)
    sheet = book[sheetname]
    result_list = [] # エクセルの結果を入れるリスト

    for line in sheet.iter_rows(min_row=min_row, min_col=min_col):
        value = []
        for item in line:
            value.append(item.value)
        result_list.append(value)
    # 保存
    book.save(filepath)
    # 終了
    book.close()

    return result_list

def excel_write(filepath, write_list, sheetname='Sheet1', min_row=0, min_col=0):
    if os.path.isfile(filepath) == True:
        exist_file_write(filepath, write_list, sheetname, min_row=0, min_col=0)
    else:
        new_file_write(filepath, write_list, sheetname, min_row=0, min_col=0)

def new_file_write(filepath, write_list, sheetname='Sheet1', min_row=0, min_col=0):
    wb = openpyxl.Workbook()
    sheet = wb.active
    row = min_row
    # 書き込み処理
    for line in write_list:
        col = min_col
        #print(line)
        for item in line:
            sheet.cell(row = row + 1, column = col + 1).value = item
            col += 1
        row += 1
    wb.save(filepath)

def exist_file_write(filepath, write_list, sheetindex=0, min_row=0, min_col=0):
    book = openpyxl.load_workbook(filepath)
    #sheet = book["Sheet"]
    sheet = book.worksheets[0]
    for line in write_list:
        max_row = sheet.max_row
        #print(max_row)
        col = 1
        for item in line:
            sheet.cell(row = max_row + 1, column = col).value = item
            col += 1

    # 保存
    book.save(filepath)
    # 終了
    book.close()